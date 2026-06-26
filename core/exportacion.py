"""
Exportación de calificaciones a Excel.

Cada exportación genera un libro con dos hojas, reflejando las mismas
dos tablas que se ven en la pestaña "Calificaciones" de la app:
  - "RA": alumnos en filas, una columna por Resultado de Aprendizaje con
    su nota, más la nota final, su calificación cualitativa y una
    columna "RA superados" (TODOS, o FALTAN: RAx, RAy...). Se incluye
    tanto en evaluaciones parciales como en FINAL, ya que es habitual
    repartir los RA entre evaluaciones y conviene ver el seguimiento
    progresivo evaluación a evaluación, no solo al cierre de curso.
  - "Criterios": alumnos en filas, una columna por criterio individual
    (con su código RA.letra), más la nota final y su calificación.

Se usa tanto para una evaluación parcial como para FINAL.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.calificacion import calificacion_cualitativa, color_hex_nota
from core.database import BaseDatosModulo, Evaluacion, Modulo

FUENTE_CABECERA = Font(bold=True, color="FFFFFF")
RELLENO_CABECERA = PatternFill("solid", start_color="E22B10")
FUENTE_NOTA_FINAL = Font(bold=True)
RELLENO_SIN_NOTA = PatternFill("solid", start_color="F2F2F2")
RELLENO_RA_TODOS_SUPERADOS = PatternFill("solid", start_color="D7F2E3")
RELLENO_RA_FALTAN = PatternFill("solid", start_color="FBE3D6")
ALINEACION_CENTRO = Alignment(horizontal="center", vertical="center")

UMBRAL_RA_SUPERADO = 5.0


def _texto_y_relleno_nota(valor: float | None) -> tuple[str, PatternFill | None]:
    if valor is None:
        return "", RELLENO_SIN_NOTA
    texto = f"{valor:.2f}".replace(".", ",")
    color_hex = color_hex_nota(valor)
    relleno = PatternFill("solid", start_color=color_hex) if color_hex else None
    return texto, relleno


def _textos_y_relleno_nota_final(valor: float | None) -> tuple[str, str, PatternFill | None]:
    if valor is None:
        return "", "", RELLENO_SIN_NOTA
    texto_numero = f"{valor:.2f}".replace(".", ",")
    texto_letra = calificacion_cualitativa(valor)
    color_hex = color_hex_nota(valor)
    relleno = PatternFill("solid", start_color=color_hex) if color_hex else None
    return texto_numero, texto_letra, relleno


def _datos_alumnos_y_notas(
    base_datos: BaseDatosModulo,
    modulo: Modulo,
    incluir_solo_evaluables: Evaluacion | None,
) -> tuple[list, dict, dict, dict]:
    if incluir_solo_evaluables is not None:
        vista = base_datos.listar_alumnos_para_evaluacion(modulo.id, incluir_solo_evaluables.orden)
        alumnos = [alumno for alumno, evaluable in vista if evaluable]
        notas_criterio = base_datos.calcular_notas_criterios_evaluacion(
            incluir_solo_evaluables.id, modulo.id
        )
        notas_ra = base_datos.calcular_notas_ra_evaluacion(incluir_solo_evaluables.id, modulo.id)
        notas_modulo = base_datos.calcular_notas_modulo_evaluacion(
            incluir_solo_evaluables.id, modulo.id
        )
    else:
        alumnos = base_datos.listar_alumnos(modulo.id)
        notas_criterio = base_datos.calcular_notas_criterios_final(modulo.id)
        notas_ra = base_datos.calcular_notas_ra_final(modulo.id)
        notas_modulo = base_datos.calcular_notas_modulo_final(modulo.id)

    return alumnos, notas_criterio, notas_ra, notas_modulo


def _anadir_hoja_ra(
    libro: Workbook,
    base_datos: BaseDatosModulo,
    modulo: Modulo,
    alumnos: list,
    notas_ra: dict,
    notas_modulo: dict,
    incluir_columna_ra_superados: bool,
) -> None:
    ras = base_datos.listar_ra(modulo.id)
    hoja = libro.create_sheet("RA")

    encabezados = ["Apellidos", "Nombre"] + [f"RA{ra.numero}" for ra in ras] + ["NOTA FINAL", "CALIFICACIÓN"]
    if incluir_columna_ra_superados:
        encabezados.append("RA SUPERADOS")

    for columna, titulo in enumerate(encabezados, start=1):
        celda = hoja.cell(row=1, column=columna, value=titulo)
        celda.font = FUENTE_CABECERA
        celda.fill = RELLENO_CABECERA
        celda.alignment = ALINEACION_CENTRO

    col_final_numero = 3 + len(ras)
    col_final_letra = col_final_numero + 1
    col_ra_superados = col_final_letra + 1

    for fila_indice, alumno in enumerate(alumnos, start=2):
        hoja.cell(row=fila_indice, column=1, value=alumno.apellidos)
        hoja.cell(row=fila_indice, column=2, value=alumno.nombre)

        ras_no_superados = []
        for col_indice, ra in enumerate(ras, start=3):
            valor = notas_ra.get((ra.id, alumno.id))
            texto, relleno = _texto_y_relleno_nota(valor)
            celda = hoja.cell(row=fila_indice, column=col_indice, value=texto)
            celda.alignment = ALINEACION_CENTRO
            if relleno is not None:
                celda.fill = relleno
            if valor is None or valor < UMBRAL_RA_SUPERADO:
                ras_no_superados.append(ra.numero)

        valor_final = notas_modulo.get(alumno.id)
        texto_numero, texto_letra, relleno_final = _textos_y_relleno_nota_final(valor_final)
        celda_numero = hoja.cell(row=fila_indice, column=col_final_numero, value=texto_numero)
        celda_letra = hoja.cell(row=fila_indice, column=col_final_letra, value=texto_letra)
        for celda_f in (celda_numero, celda_letra):
            celda_f.font = FUENTE_NOTA_FINAL
            celda_f.alignment = ALINEACION_CENTRO
            if relleno_final is not None:
                celda_f.fill = relleno_final

        if incluir_columna_ra_superados:
            if ras_no_superados:
                texto_superados = "FALTAN: " + ", ".join(f"RA{n}" for n in ras_no_superados)
                relleno_superados = RELLENO_RA_FALTAN
            else:
                texto_superados = "TODOS"
                relleno_superados = RELLENO_RA_TODOS_SUPERADOS
            celda_superados = hoja.cell(row=fila_indice, column=col_ra_superados, value=texto_superados)
            celda_superados.font = FUENTE_NOTA_FINAL
            celda_superados.alignment = ALINEACION_CENTRO
            celda_superados.fill = relleno_superados

    hoja.column_dimensions["A"].width = 22
    hoja.column_dimensions["B"].width = 18
    ultima_columna = col_ra_superados if incluir_columna_ra_superados else col_final_letra
    for col_indice in range(3, ultima_columna + 1):
        ancho = 30 if (incluir_columna_ra_superados and col_indice == col_ra_superados) else 16
        hoja.column_dimensions[get_column_letter(col_indice)].width = ancho
    hoja.freeze_panes = "C2"


def _anadir_hoja_criterios(
    libro: Workbook,
    base_datos: BaseDatosModulo,
    modulo: Modulo,
    alumnos: list,
    notas_criterio: dict,
    notas_modulo: dict,
) -> None:
    criterios_con_ra = base_datos.listar_criterios_de_modulo(modulo.id)
    codigos = [base_datos.codigo_criterio(ra, criterio) for ra, criterio in criterios_con_ra]
    hoja = libro.create_sheet("Criterios")

    encabezados = ["Apellidos", "Nombre"] + [""] * len(criterios_con_ra) + ["NOTA FINAL", "CALIFICACIÓN"]
    col_final_numero = 3 + len(criterios_con_ra)
    col_final_letra = col_final_numero + 1

    for columna, titulo in enumerate(encabezados, start=1):
        celda = hoja.cell(row=1, column=columna, value=titulo)
        celda.font = FUENTE_CABECERA
        celda.fill = RELLENO_CABECERA
        celda.alignment = ALINEACION_CENTRO
    hoja.cell(row=1, column=1, value="Apellidos")
    hoja.cell(row=1, column=2, value="Nombre")

    fila_codigo = 2
    fila_peso = 3
    hoja.cell(row=fila_codigo, column=1, value="Criterio")
    hoja.cell(row=fila_peso, column=1, value="Peso en su RA")
    for fila_etiqueta in (fila_codigo, fila_peso):
        celda_et = hoja.cell(row=fila_etiqueta, column=1)
        celda_et.font = FUENTE_NOTA_FINAL
    for col_indice, (codigo, (_ra, criterio)) in enumerate(zip(codigos, criterios_con_ra), start=3):
        celda_codigo = hoja.cell(row=fila_codigo, column=col_indice, value=codigo)
        celda_codigo.font = FUENTE_NOTA_FINAL
        celda_codigo.alignment = ALINEACION_CENTRO
        celda_peso = hoja.cell(row=fila_peso, column=col_indice, value=criterio.peso)
        celda_peso.alignment = ALINEACION_CENTRO

    fila_primer_alumno = 4
    for indice_alumno, alumno in enumerate(alumnos):
        fila_indice = fila_primer_alumno + indice_alumno
        hoja.cell(row=fila_indice, column=1, value=alumno.apellidos)
        hoja.cell(row=fila_indice, column=2, value=alumno.nombre)
        for col_indice, (_ra, criterio) in enumerate(criterios_con_ra, start=3):
            valor = notas_criterio.get((criterio.id, alumno.id))
            texto, relleno = _texto_y_relleno_nota(valor)
            celda = hoja.cell(row=fila_indice, column=col_indice, value=texto)
            celda.alignment = ALINEACION_CENTRO
            if relleno is not None:
                celda.fill = relleno

        valor_final = notas_modulo.get(alumno.id)
        texto_numero, texto_letra, relleno_final = _textos_y_relleno_nota_final(valor_final)
        celda_numero = hoja.cell(row=fila_indice, column=col_final_numero, value=texto_numero)
        celda_letra = hoja.cell(row=fila_indice, column=col_final_letra, value=texto_letra)
        for celda_f in (celda_numero, celda_letra):
            celda_f.font = FUENTE_NOTA_FINAL
            celda_f.alignment = ALINEACION_CENTRO
            if relleno_final is not None:
                celda_f.fill = relleno_final

    hoja.column_dimensions["A"].width = 22
    hoja.column_dimensions["B"].width = 18
    for col_indice in range(3, col_final_letra + 1):
        hoja.column_dimensions[get_column_letter(col_indice)].width = 16
    hoja.freeze_panes = f"C{fila_primer_alumno}"


def _construir_libro(
    base_datos: BaseDatosModulo,
    modulo: Modulo,
    incluir_solo_evaluables: Evaluacion | None,
) -> Workbook:
    alumnos, notas_criterio, notas_ra, notas_modulo = _datos_alumnos_y_notas(
        base_datos, modulo, incluir_solo_evaluables
    )

    libro = Workbook()
    libro.remove(libro.active)

    _anadir_hoja_ra(libro, base_datos, modulo, alumnos, notas_ra, notas_modulo, incluir_columna_ra_superados=True)
    _anadir_hoja_criterios(libro, base_datos, modulo, alumnos, notas_criterio, notas_modulo)

    return libro


def exportar_evaluacion(
    base_datos: BaseDatosModulo, modulo: Modulo, evaluacion: Evaluacion, ruta_destino: str | Path
) -> Path:
    """Exporta las calificaciones de una evaluación parcial."""
    libro = _construir_libro(base_datos, modulo, incluir_solo_evaluables=evaluacion)
    ruta_destino = Path(ruta_destino)
    libro.save(ruta_destino)
    return ruta_destino


def exportar_final(base_datos: BaseDatosModulo, modulo: Modulo, ruta_destino: str | Path) -> Path:
    """Exporta las calificaciones de la evaluación FINAL (todo el curso)."""
    libro = _construir_libro(base_datos, modulo, incluir_solo_evaluables=None)
    ruta_destino = Path(ruta_destino)
    libro.save(ruta_destino)
    return ruta_destino
