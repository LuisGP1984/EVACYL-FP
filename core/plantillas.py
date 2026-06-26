"""
Generación de plantillas Excel de ejemplo, para que el docente sepa
exactamente qué columnas debe rellenar antes de usar "Importar archivo
.xlsx…" en Alumnos o en Criterios.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

FUENTE_CABECERA = Font(bold=True, color="FFFFFF")
RELLENO_CABECERA = PatternFill("solid", start_color="2E7D4F")
ALINEACION_CENTRO = Alignment(horizontal="center", vertical="center")


def generar_plantilla_alumnos(ruta_destino: str | Path) -> Path:
    """Genera un .xlsx de ejemplo con el formato esperado para importar
    alumnado: dos columnas, Apellidos y Nombre, con un par de filas de
    ejemplo (que el docente debe sustituir por su alumnado real).
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Alumnos"

    encabezados = ["Apellidos", "Nombre"]
    for columna, titulo in enumerate(encabezados, start=1):
        celda = hoja.cell(row=1, column=columna, value=titulo)
        celda.font = FUENTE_CABECERA
        celda.fill = RELLENO_CABECERA
        celda.alignment = ALINEACION_CENTRO

    ejemplos = [
        ("García Pérez", "Ana"),
        ("López Ruiz", "Luis"),
        ("Martínez Sánchez", "Eva"),
    ]
    for fila, (apellidos, nombre) in enumerate(ejemplos, start=2):
        hoja.cell(row=fila, column=1, value=apellidos)
        hoja.cell(row=fila, column=2, value=nombre)

    hoja.column_dimensions["A"].width = 28
    hoja.column_dimensions["B"].width = 20

    ruta_destino = Path(ruta_destino)
    libro.save(ruta_destino)
    return ruta_destino


def generar_plantilla_criterios(ruta_destino: str | Path) -> Path:
    """Genera un .xlsx de ejemplo con el formato esperado para importar
    criterios de evaluación: dos columnas, Código y Peso, con un par de
    filas de ejemplo siguiendo la nomenclatura LOMLOE (competencia.criterio).
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Criterios"

    encabezados = ["Código", "Peso"]
    for columna, titulo in enumerate(encabezados, start=1):
        celda = hoja.cell(row=1, column=columna, value=titulo)
        celda.font = FUENTE_CABECERA
        celda.fill = RELLENO_CABECERA
        celda.alignment = ALINEACION_CENTRO

    ejemplos = [
        ("1.1", 1),
        ("1.2", 1),
        ("2.1", 2),
    ]
    for fila, (codigo, peso) in enumerate(ejemplos, start=2):
        hoja.cell(row=fila, column=1, value=codigo)
        hoja.cell(row=fila, column=2, value=peso)

    hoja.column_dimensions["A"].width = 18
    hoja.column_dimensions["B"].width = 14

    ruta_destino = Path(ruta_destino)
    libro.save(ruta_destino)
    return ruta_destino
