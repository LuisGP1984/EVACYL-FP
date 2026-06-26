"""
Utilidades para traer datos desde Excel hacia las tablas de la aplicación,
de dos formas:

  - Pegado (Ctrl+V): el docente copia celdas en Excel y las pega directamente
    en la tabla de la app. Qt entrega ese contenido como texto con
    tabulaciones entre columnas y saltos de línea entre filas (es el
    formato estándar al copiar celdas de Excel), así que basta con
    interpretarlo como tal.

  - Importar archivo .xlsx: se abre el archivo con openpyxl y se leen
    las columnas indicadas.
"""

from __future__ import annotations

from openpyxl import load_workbook


def filas_desde_texto_pegado(texto: str) -> list[list[str]]:
    """Convierte el texto pegado (formato portapapeles de Excel) en una
    lista de filas, cada una como lista de columnas (texto plano).
    Ignora filas completamente vacías.
    """
    filas = []
    for linea in texto.splitlines():
        if linea.strip() == "":
            continue
        columnas = linea.split("\t")
        filas.append([c.strip() for c in columnas])
    return filas


def filas_desde_excel(ruta_archivo: str, nombre_hoja: str | None = None) -> list[list[str]]:
    """Lee una hoja de un archivo .xlsx y devuelve sus filas como texto.
    Si no se indica nombre_hoja, se usa la primera hoja del libro.
    Las celdas vacías se devuelven como cadena vacía.
    """
    libro = load_workbook(ruta_archivo, data_only=True, read_only=True)
    hoja = libro[nombre_hoja] if nombre_hoja else libro[libro.sheetnames[0]]

    filas = []
    for fila_celdas in hoja.iter_rows(values_only=True):
        if all(valor is None or str(valor).strip() == "" for valor in fila_celdas):
            continue
        fila_textos = [("" if valor is None else str(valor).strip()) for valor in fila_celdas]
        filas.append(fila_textos)
    libro.close()
    return filas


def normalizar_filas_alumnos(filas: list[list[str]]) -> list[tuple[str, str]]:
    """A partir de filas crudas (de pegado o de Excel), devuelve tuplas
    (apellidos, nombre). Acepta filas de 1 o 2 columnas:
      - 2 columnas: (apellidos, nombre) directamente.
      - 1 columna: se interpreta como "Apellidos, Nombre" si hay coma,
        o como un único campo de apellidos si no la hay.
    Se descarta automáticamente una posible fila de cabecera
    (ej. "Apellidos" / "Nombre") si coincide por contenido.
    """
    resultado = []
    for fila in filas:
        fila = [c for c in fila if c != ""] if len(fila) > 2 else fila
        primera_celda_minuscula = fila[0].strip().lower() if fila else ""
        if primera_celda_minuscula in ("apellidos", "apellido y nombre", "alumno", "alumno/a"):
            continue  # probablemente es la fila de cabecera
        if len(fila) >= 2:
            apellidos, nombre = fila[0], fila[1]
        elif len(fila) == 1 and "," in fila[0]:
            apellidos, nombre = (parte.strip() for parte in fila[0].split(",", 1))
        elif len(fila) == 1:
            apellidos, nombre = fila[0], ""
        else:
            continue
        if apellidos or nombre:
            resultado.append((apellidos, nombre))
    return resultado


def normalizar_filas_criterios(filas: list[list[str]]) -> list[tuple[str, float]]:
    """A partir de filas crudas (de pegado o de Excel), devuelve tuplas
    (codigo, peso). Si no se indica peso, se usa 1.0 por defecto.
    Descarta una posible fila de cabecera.
    """
    resultado = []
    for fila in filas:
        primera_celda_minuscula = fila[0].strip().lower() if fila else ""
        if primera_celda_minuscula in ("codigo", "código", "criterio"):
            continue
        if not fila or not fila[0]:
            continue
        codigo = fila[0]
        peso = 1.0
        if len(fila) >= 2 and fila[1]:
            try:
                peso = float(fila[1].replace(",", "."))
            except ValueError:
                peso = 1.0
        resultado.append((codigo, peso))
    return resultado
